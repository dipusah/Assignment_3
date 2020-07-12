import os
import xlrd
import xlsxwriter
import openpyxl
from openpyxl import Workbook


# Greeter is a terminal application that greets old friends warmly,
#   and remembers new friends.


### FUNCTIONS ###
def update_register_student_details():
    path = "student_info.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    # from the active attribute
    sheet_obj = wb_obj.active
    i = int(input("enter row:"))
    j = int(input("enter column:"))

    row_and_column = sheet_obj.cell(row=i, column=j)
    cell_value = input("input cell value you want to update:")
    row_and_column.value = cell_value
    wb_obj.save("updated_info.xlsx")


def register_student_details():
    print("Name ", "|", "Coures enrolled", '|', "ammount_paid", "|", "amount_remaining")
    path = 'student_info.xlsx'
    inputWorkbook = xlrd.open_workbook(path)
    inputWorksheet = inputWorkbook.sheet_by_index(0)
    for i in range(1, inputWorksheet.nrows):
        print(inputWorksheet.cell_value(i, 0), '|', inputWorksheet.cell_value(i, 1), "|",
              inputWorksheet.cell_value(i, 2), "|", inputWorksheet.cell_value(i, 3))

def student_information_xlsx_file():
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook('student_info.xlsx')
    worksheet = workbook.add_worksheet()

    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})

    # Add a number format for cells with money.
    money = workbook.add_format({'num_format': '$#,##0'})

    # Write some data headers.
    worksheet.write('A1', 'Name', bold)
    worksheet.write('B1', 'Courses', bold)
    worksheet.write('C1', 'Amount_paid', bold)
    worksheet.write('D1', 'Ammount_remainng', bold)

    # Some data we want to write to the worksheet.
    expenses = (
        ['Dipak', 'NETWORK ENGINEER COURSE (1 YEAR)', 1000, 1000],
        ['Hari', 'IT ENGINEER COURSE (6 MONTHS)', 1000, 1000],
        ['Shyam', 'SOFTWARE DEVELOPER (1 YEAR)', 1000, 1000],
        ['Gita', 'SOFTWARE DEVELOPER (6 MONTHS)', 1000, 1000],
    )

    # Start from the first cell below the headers.
    row = 1
    col = 0

    # Iterate over the data and write it out row by row.
    for nam, course, amt_p, amt_r in (expenses):
        worksheet.write(row, col, nam)
        worksheet.write(row, col + 1, course)
        worksheet.write(row, col + 2, amt_p, money)
        worksheet.write(row, col + 3, amt_r, money)
        row += 1

    workbook.close()

def display_title_bar():
    # Clears the terminal screen, and displays a title bar.
    os.system('clear')

    print("\t**********************************************")
    print("\t***          IT ACADEMY !                  ***")
    print("\t**********************************************")


def get_user_choice():
    # Let users know what they can do.
    print("\n[1] program course of study.")
    print("[2] Inquiry about the course of study.")
    print("[3] student registration fee.")
    print("[4] student information with their payments and remaining payments.")
    print("[5] update student information.")
    print("[6] Deletion of student program if he/she left the program.")
    print("[7] return the deposit amount after completion of course.")
    print("[q] Quit.")
def enter_choice():
    return input("What would you like to do? enter your choice = > ")


def perform_more_action():
    option = input("would you like to perform more action(1-'yes'/0-'No'):->")
    if option == '1':
        get_user_choice()
        # Respond to the user's choice.

        display_title_bar()

    else:
        print("Thanks for visiting")
        exit()





### MAIN PROGRAM ###

# Set up a loop where users can choose what they'd like to do.
choice = ''
display_title_bar()

while choice != 'q':

    path = 'courses_detai.xlsx'
    inputWorkbook = xlrd.open_workbook(path)
    inputWorksheet = inputWorkbook.sheet_by_index(0)

    get_user_choice()
    choice=enter_choice()

    # Respond to the user's choice.
    display_title_bar()
    if choice == '1':
        print('**********************************************************')

        for i in range(1, inputWorksheet.nrows):
            print(inputWorksheet.cell_value(i, 0))

        print('*********************************************************')
        perform_more_action()

    elif choice == '2':
        print("\ncourses detail")

        for i in range(1, inputWorksheet.nrows):
            print(inputWorksheet.cell_value(i, 0), "==>>", inputWorksheet.cell_value(i, 1))
        perform_more_action()


    elif choice == '3':
        print("student registration fee each 1000 for installment :- ", 2000)
        print("Student are allow to pay  amount in two installment 1000 each: ")
        print("student details are taken in student_info file: ")
        student_information_xlsx_file()
        perform_more_action()

    elif choice == '4':
        register_student_details()
        perform_more_action()

    elif choice == '5':
        print("updated new information")
        register_student_details()
        print("look above table and insert row and column value for update cell value:")
        update_register_student_details()
        perform_more_action()

    elif choice == '6':
        print("delete student information who left program for that insert row number of student who left:")
        register_student_details()

        path = "student_info.xlsx"
        wb_obj = openpyxl.load_workbook(path.strip())
        # from the active attribute
        sheet_obj = wb_obj.active
        n=int(input("enter row you want to delete:"))
        sheet_obj.delete_rows(n)
        wb_obj.save("new_student_info.xlsx") #deleted student remove from file student info and remaining data save into new_student_info file
        perform_more_action()

    elif choice == '7':

        student='course_complete'
        if (student=='course_complete'):
            print("return the deposited 2000")
        perform_more_action()

    elif choice == 'q':
        exit()

    else:
        print("\nI didn't understand that choice.\n")
