"""import xlsxwriter
import openpyxl

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook('Expenses02.xlsx')
worksheet = workbook.add_worksheet()

# Add a bold format to use to highlight cells.
bold = workbook.add_format({'bold': True})

# Add a number format for cells with money.
money = workbook.add_format({'num_format': '$#,##0'})

# Write some data headers.
worksheet.write('A1', 'Name', bold)
worksheet.write('B1', 'Courses', bold)
worksheet.write('C1','Amount_paid',bold)
worksheet.write('D1','Ammount_remainng',bold)


# Some data we want to write to the worksheet.
expenses = (
    ['Dipak', 'NETWORK ENGINEER COURSE (1 YEAR)',1000,1000],
    ['Hari','IT ENGINEER COURSE (6 MONTHS)',1000,1000],
    ['Shyam','SOFTWARE DEVELOPER (1 YEAR)', 1000, 1000],
    ['Gita','SOFTWARE DEVELOPER (6 MONTHS)',1000,1000],
)

# Start from the first cell below the headers.
row = 1
col = 0

# Iterate over the data and write it out row by row.
for nam, course, amt_p,amt_r in (expenses):
    worksheet.write(row, col, nam)
    worksheet.write(row, col+1, course)
    worksheet.write(row, col + 2, amt_p, money)
    worksheet.write(row, col+3, amt_r,money)
    row += 1

workbook.close()"""

import openpyxl
from openpyxl import Workbook

path = "Expenses02.xlsx"
wb_obj = openpyxl.load_workbook(path.strip())
# from the active attribute
sheet_obj = wb_obj.active
sheet_obj.delete_rows(2)
wb_obj.save("del.xlsx")

"""i= int(input("enter row:"))
j=int(input("enter column:"))

row_and_column=sheet_obj.cell(row=i,column=j)
cell_value=input("input cell value you want to update:")
row_and_column.value=cell_value
wb_obj.save("Expenses02.xlsx")"""

